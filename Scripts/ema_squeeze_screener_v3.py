"""
Weekly EMA Squeeze & Crossover Screener for Nifty 500
Signals: 10 WEMA, 30 WEMA, 40 WEMA squeezing + 10 WEMA crossing above 30/40 WEMA
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import os, warnings
warnings.filterwarnings("ignore")

# ── Stock Universe ────────────────────────────────────────────────────────────
NIFTY500_STOCKS = [
    "RELIANCE","TCS","HDFCBANK","BHARTIARTL","ICICIBANK","INFOSYS","SBIN",
    "HINDUNILVR","ITC","LT","KOTAKBANK","AXISBANK","BAJFINANCE","MARUTI",
    "NTPC","ONGC","POWERGRID","ULTRACEMCO","TATAMOTORS","WIPRO","ADANIENT",
    "ADANIPORTS","BAJAJFINSV","HCLTECH","SUNPHARMA","TITAN","TECHM","NESTLEIND",
    "COALINDIA","GRASIM","INDUSINDBK","HINDALCO","JSWSTEEL","TATACONSUM",
    "CIPLA","DIVISLAB","DRREDDY","EICHERMOT","BRITANNIA","APOLLOHOSP",
    "BAJAJ_AUTO","BPCL","HEROMOTOCO","TATAPOWER","HAVELLS","PIDILITEIND",
    "MUTHOOTFIN","LUPIN","BERGEPAINT","GODREJCP","MOTHERSON","DABUR",
    "MARICO","COLPAL","TORNTPHARM","BOSCHLTD","SIEMENS","ABB","BHEL",
    "SAIL","NMDC","VEDL","HINDPETRO","IOC","GAIL","PETRONET","MGL",
    "IGL","ATGL","ADANIGREEN","ADANITRANS","ADANIGAS","TATACOMM",
    "MPHASIS","LTIM","PERSISTENT","COFORGE","KPITTECH","ZOMATO","NYKAA",
    "PAYTM","POLICYBZR","DELHIVERY","IRCTC","CONCOR","GMRINFRA","ADANIPOWER",
    "TRENT","DMART","VSTIND","PAGEIND","RELAXO","BATAINDIA","WHIRLPOOL",
    "VOLTAS","BLUESTARCO","CROMPTON","POLYCAB","KECL","KALPATPOWR",
    "CUMMINSIND","THERMAX","BEL","HAL","DRDO","BHEL","BEML","MIDHANI",
    "MFSL","MAXHEALTH","FORTIS","NARAYANA","THYROCARE","METROPOLIS",
    "LALPATHLAB","VIJAYABANK","FEDERALBNK","IDFCFIRSTB","RBLBANK",
    "BANDHANBNK","AUBANK","UJJIVANSFB","EQUITASBNK","ESAFSFB",
    "CHOLAFIN","BAJAJHLDNG","ICICIGI","HDFCLIFE","SBILIFE","LICI",
    "STARHEALTH","NIACL","UCOBANK","CENTRALBK","MAHABANK","BANKINDIA",
    "UNIONBANK","PNB","CANBK","IOB","INDIANB","BANKBARODA",
    "ASHOKLEY","TVSMOTOR","MAHINDRA","ESCORTS","SWARAJENG","FORCEMOT",
    "TATACHEM","PIDILITIND","AARTIIND","SUDARSCHEM","NAVINFLUOR",
    "DEEPAKNTR","VINATIORGA","BALAMINES","ALKYLAMINE","CLEAN","FINEORG",
    "SRF","GARFIBRES","RAYMOND","ARVIND","TRIDENT","WELSPUN","HIMATSEIDE",
    "BALRAMCHIN","RENUKA","TRIVENI","DHANUKA","BAYER","PIIND","RALLIS",
    "ASTRAL","SUPREME","APCOTEX","GHCL","VINDHYATEL","HFCL","STLTECH",
    "TEJASNET","GTLINFRA","RAILTEL","IRFC","RECLTD","PFC","HUDCO",
    "NHPC","SJVN","NLCINDIA","CESC","TORNTPOWER","JINDALSTEEL","RATNAMANI",
    "WELCORP","MAHARASTRA","APL","JSWENERGY","GREENPANEL","CENTURYPLY",
    "GREENPLY","PLYBOARD","SBIN","CANFINHOME","LICHSGFIN","PNBHOUSING",
    "AAVAS","HOMEFIRST","APTUS","REPCO","GRUH","INDIABULL","IBULHSGFIN",
    "CHOLAHLDNG","SUNDARMFIN","M_MFIN","MAHFIN","SHRIRAMFIN","MANAPPURAM",
    "CREDITACC","SPANDANA","ARMANFIN","FUSION","UGROCAP",
    "ZYDUSLIFE","ALKEM","GLENMARK","IPCA","AJANTPHARM","NATCOPHARM",
    "GRANULES","SUVEN","LAURUSLABS","SOLARA","SEQUENT","IOLCP","NEULANDLAB",
    "INGERRAND","KENNAMETAL","ELGIEQUIP","GRINDWELL","CARBORUNIV","AIA",
    "SCHAEFFLER","SKFIND","TIMKEN","NAUKRI","JUSTDIAL","INFOEDGE","ZAUBACORP",
    "CARTRADE","EASEMYTRIP","IXIGO","YATRA","MAKEMYTRIP",
    "AFFLE","TANLA","ONMOBILE","ROUTE","GTPL","INDIAMART",
    "TATAELXSI","LTTS","CYIENT","HEXAWARE","NIITTECH","RAMSYSTEMS",
    "SYMPHONY","WHIRLPOOL","AMBER","DIXON","VGUARD","ORIENTELEC",
    "KNRCON","NCC","PNCINFRA","HGINFRA","GPPL","ADANIPORTS",
    "JINDALSAW","MAHSEAMLES","RATNAMANI","WELCORP","TITAGARH","TEXRAIL",
    "IRCON","RITES","NBCC","WABCOINDIA","SUNDRMFAST","ENDURANCE",
    "SUPRAJIT","GABRIEL","MINDA","UNOMINDA","SAMVARDHANA","SANSERA",
    "CRAFTSMAN","METALFORGE","BHARAT","EXIDEIND","AMARAJABAT","HBLPOWER",
    "TIDEWATER","GESHIP","COCHINSHIP","MAZAGON","GDL","ALLCARGO",
    "TCI","MAHLOG","BLUEDART","GATI","XPRESSBEES",
    "TASTYBITE","VARUN","GLOBUSSPR","PATANJALI","EMAMILTD","JYOTHYLAB",
    "VENKEYS","ZYDUSWELL","HONAUT","3MINDIA","PFIZER","ABBOTINDIA",
    "SANOFI","GLAXO","ASTRAZEN","NOVARTIS","BIOCON","STRIDES",
    "IPCALAB","MARKSANS","OPTIEMUS","AMBER","KAYNES","SYRMA",
    "ELCID","FIEM","SUPRAJIT","SANDHAR","MINDA","EXIDEIND"
]

# Remove duplicates
NIFTY500_STOCKS = list(dict.fromkeys(NIFTY500_STOCKS))

EMA_SHORT  = 10
EMA_MID    = 30
EMA_LONG   = 40
SQUEEZE_PCT = 5.0   # EMAs within 5% of each other = squeezing (covers high beta stocks)
LOOKBACK_WEEKS = 60  # ~15 months of weekly data

# ── EMA Calculation ───────────────────────────────────────────────────────────
def calc_ema(series, span):
    return series.ewm(span=span, adjust=False).mean()

def calc_rsi(series, period=14):
    delta = series.diff()
    gain  = delta.clip(lower=0).rolling(period).mean()
    loss  = (-delta.clip(upper=0)).rolling(period).mean()
    rs    = gain / loss
    return 100 - (100 / (1 + rs))

def calc_adx(df, period=14):
    high, low, close = df["High"], df["Low"], df["Close"]
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    atr   = tr.ewm(span=period, adjust=False).mean()
    up    = high.diff()
    down  = -low.diff()
    dm_p  = pd.Series(np.where((up > down) & (up > 0), up, 0), index=df.index)
    dm_m  = pd.Series(np.where((down > up) & (down > 0), down, 0), index=df.index)
    di_p  = 100 * dm_p.ewm(span=period, adjust=False).mean() / atr
    di_m  = 100 * dm_m.ewm(span=period, adjust=False).mean() / atr
    dx    = 100 * (di_p - di_m).abs() / (di_p + di_m)
    return dx.ewm(span=period, adjust=False).mean()

# ── Squeeze Detection ─────────────────────────────────────────────────────────
def is_squeezing(e10, e30, e40, pct=SQUEEZE_PCT):
    """All 3 EMAs within pct% of each other"""
    hi = max(e10, e30, e40)
    lo = min(e10, e30, e40)
    return ((hi - lo) / lo * 100) <= pct

# ── Main Screening ────────────────────────────────────────────────────────────
def screen_stock(symbol):
    try:
        ticker = symbol + ".NS"
        df = yf.download(ticker, period="15mo", interval="1wk",
                         progress=False, auto_adjust=True)
        if df.empty or len(df) < LOOKBACK_WEEKS:
            return None

        df = df.copy()
        df["EMA10"] = calc_ema(df["Close"], EMA_SHORT)
        df["EMA30"] = calc_ema(df["Close"], EMA_MID)
        df["EMA40"] = calc_ema(df["Close"], EMA_LONG)
        df["RSI"]   = calc_rsi(df["Close"])
        df["ADX"]   = calc_adx(df)
        df["Vol20"] = df["Volume"].rolling(20).mean()

        cur = df.iloc[-1]
        prv = df.iloc[-2]

        e10_c, e30_c, e40_c = cur["EMA10"], cur["EMA30"], cur["EMA40"]
        e10_p, e30_p, e40_p = prv["EMA10"], prv["EMA30"], prv["EMA40"]
        close = float(cur["Close"])

        # ── Signal Detection ──────────────────────────────────────────────────
        # MANDATORY FILTER: 30W EMA must be above 40W EMA
        # This confirms medium-term trend is rising above long-term baseline
        ema30_above_ema40 = e30_c > e40_c
        if not ema30_above_ema40:
            return None

        squeeze_now  = is_squeezing(e10_c, e30_c, e40_c)
        squeeze_prev = is_squeezing(e10_p, e30_p, e40_p)

        # ── All shorter WEMA crossing longer WEMA ─────────────────────────────
        # Pair 1: 10W crossing above 30W  (short × mid)
        cross_10x30 = (e10_p < e30_p) and (e10_c > e30_c)
        # Pair 2: 10W crossing above 40W  (short × long)
        cross_10x40 = (e10_p < e40_p) and (e10_c > e40_c)
        # Pair 3: 30W crossing above 40W  (mid × long) — strongest structural shift
        cross_30x40 = (e30_p < e40_p) and (e30_c > e40_c)

        any_cross  = cross_10x30 or cross_10x40 or cross_30x40
        # Perfect bull stack: 10W > 30W > 40W
        full_stack = (e10_c > e30_c) and (e30_c > e40_c)

        # EMA spread % (lower = tighter squeeze)
        hi = max(e10_c, e30_c, e40_c)
        lo = min(e10_c, e30_c, e40_c)
        spread_pct = (hi - lo) / lo * 100

        # Signal classification — priority 1 is best
        if (squeeze_prev or squeeze_now) and any_cross and full_stack:
            signal = "🔥 SQUEEZE + FULL STACK"
            priority = 1
        elif (squeeze_prev or squeeze_now) and any_cross:
            signal = "🔥 SQUEEZE + CROSSOVER"
            priority = 2
        elif squeeze_now and full_stack:
            signal = "⚡ SQUEEZE BREAKOUT"
            priority = 3
        elif cross_10x30 and cross_10x40 and cross_30x40:
            signal = "✅ ALL 3 CROSSED"
            priority = 4
        elif cross_30x40 and cross_10x30:
            signal = "📈 30x40 + 10x30 CROSS"
            priority = 5
        elif cross_30x40 and cross_10x40:
            signal = "📈 30x40 + 10x40 CROSS"
            priority = 5
        elif cross_30x40:
            signal = "📈 30 X 40 CROSS"
            priority = 6
        elif cross_10x30 and cross_10x40:
            signal = "📈 10x30 + 10x40 CROSS"
            priority = 7
        elif cross_10x30:
            signal = "📈 10 X 30 CROSS"
            priority = 8
        elif cross_10x40:
            signal = "📈 10 X 40 CROSS"
            priority = 9
        elif squeeze_now:
            signal = "🔄 SQUEEZING"
            priority = 10
        else:
            return None   # Skip non-signals

        # Above 200-week EMA filter
        ema200 = calc_ema(df["Close"], 200).iloc[-1]
        trend  = "BULL" if close > ema200 else "BEAR"

        # 52-week high
        w52_high = df["High"].iloc[-52:].max() if len(df) >= 52 else df["High"].max()
        pct_from_high = (close - w52_high) / w52_high * 100

        vol_ratio = float(cur["Volume"]) / float(cur["Vol20"]) if cur["Vol20"] > 0 else 0

        return {
            "Symbol"         : symbol,
            "Signal"         : signal,
            "Priority"       : priority,
            "Close"          : round(close, 2),
            "EMA10W"         : round(e10_c, 2),
            "EMA30W"         : round(e30_c, 2),
            "EMA40W"         : round(e40_c, 2),
            "Spread%"        : round(spread_pct, 2),
            "RSI_W"          : round(float(cur["RSI"]), 1),
            "ADX_W"          : round(float(cur["ADX"]), 1),
            "Vol_Ratio"      : round(vol_ratio, 2),
            "Trend"          : trend,
            "From52wHigh%"   : round(pct_from_high, 1),
            "Full_Stack"     : "✅" if full_stack else "NO",
            "Squeeze_Now"    : "YES" if squeeze_now else "NO",
            "Cross_10x30"    : "YES" if cross_10x30 else "NO",
            "Cross_10x40"    : "YES" if cross_10x40 else "NO",
            "Cross_30x40"    : "YES" if cross_30x40 else "NO",
            "30W_above_40W"  : "✅",   # Always YES — mandatory filter
        }

    except Exception as e:
        return None

# ── Excel Report ──────────────────────────────────────────────────────────────
DARK_BG     = "1E1E2E"
HEADER_BG   = "2D2D44"
FIRE_ORANGE = "FF6B35"
ELECTRIC    = "00D4FF"
GREEN_BULL  = "00FF88"
YELLOW_SQZ  = "FFD700"
BLUE_CROSS  = "4FC3F7"
GREY_TEXT   = "B0B0C0"
WHITE       = "FFFFFF"

SIGNAL_COLORS = {
    "🔥 SQUEEZE + CROSSOVER" : ("FF6B35", "FFFFFF"),
    "⚡ SQUEEZE BREAKOUT"    : ("FFD700", "1E1E2E"),
    "✅ DOUBLE CROSSOVER"    : ("00FF88", "1E1E2E"),
    "📈 10 X 30 CROSS"       : ("4FC3F7", "1E1E2E"),
    "📈 10 X 40 CROSS"       : ("6EC6FF", "1E1E2E"),
    "🔄 SQUEEZING"           : ("9C88FF", "FFFFFF"),
}

def thin_border():
    s = Side(style="thin", color="3A3A5C")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(bold=True, size=10, color=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_excel(results, output_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Signals ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "EMA Squeeze Signals"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF6B35"

    # Title row
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = f"📊  NSE Weekly EMA Squeeze & Crossover Screener  |  Run: {datetime.now().strftime('%d %b %Y')}"
    t.font  = Font(name="Calibri", bold=True, size=14, color=FIRE_ORANGE)
    t.fill  = cell_fill(DARK_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sub-header row
    ws.merge_cells("A2:P2")
    s = ws["A2"]
    s.value = f"Signal: 10W EMA × 30W EMA × 40W EMA Squeeze + Crossover  |  Universe: Nifty 500  |  Timeframe: Weekly"
    s.font  = Font(name="Calibri", size=9, color=GREY_TEXT, italic=True)
    s.fill  = cell_fill(DARK_BG)
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Column headers
    headers = [
        "Symbol", "Signal Type", "Close ₹", "EMA 10W", "EMA 30W", "EMA 40W",
        "Spread %", "RSI (W)", "ADX (W)", "Vol Ratio",
        "Trend", "From 52W High %", "Squeeze", "10×30", "10×40", "TradingView"
    ]
    col_widths = [14, 24, 11, 11, 11, 11, 10, 9, 9, 10, 8, 15, 9, 7, 7, 18]

    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font      = hdr_font(bold=True, size=9)
        c.fill      = cell_fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 28

    # Data rows
    sorted_results = sorted(results, key=lambda x: (x["Priority"], x["Spread%"]))
    row = 4
    for r in sorted_results:
        sig   = r["Signal"]
        bg, fg = SIGNAL_COLORS.get(sig, (HEADER_BG, WHITE))
        row_bg = "252538" if row % 2 == 0 else "1E1E2E"

        values = [
            r["Symbol"], sig, r["Close"],
            r["EMA10W"], r["EMA30W"], r["EMA40W"],
            r["Spread%"], r["RSI_W"], r["ADX_W"],
            r["Vol_Ratio"], r["Trend"],
            r["From52wHigh%"],
            r["Squeeze_Now"], r["Cross_10x30"], r["Cross_10x40"],
            f"https://www.tradingview.com/chart/?symbol=NSE:{r['Symbol']}"
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font   = Font(name="Calibri", size=9, color=WHITE)
            c.fill   = cell_fill(row_bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Signal cell styling
        sig_cell = ws.cell(row=row, column=2)
        sig_cell.fill = cell_fill(bg)
        sig_cell.font = Font(name="Calibri", bold=True, size=9, color=fg)

        # Trend coloring
        t_cell = ws.cell(row=row, column=11)
        t_cell.font = Font(name="Calibri", bold=True, size=9,
                           color=GREEN_BULL if r["Trend"] == "BULL" else "FF4444")

        # RSI coloring
        rsi_cell = ws.cell(row=row, column=8)
        rsi = r["RSI_W"]
        rsi_color = "00FF88" if rsi > 60 else ("FFD700" if rsi > 45 else "FF6B6B")
        rsi_cell.font = Font(name="Calibri", size=9, color=rsi_color)

        # TradingView hyperlink
        tv_cell = ws.cell(row=row, column=16)
        tv_cell.value = f"NSE:{r['Symbol']}"
        tv_cell.hyperlink = f"https://www.tradingview.com/chart/?symbol=NSE:{r['Symbol']}"
        tv_cell.font = Font(name="Calibri", size=9, color="4FC3F7", underline="single")

        row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:P{row-1}"

    # ── Sheet 2: Priority Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("🔥 Top Picks")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "FFD700"

    top_picks = [r for r in sorted_results if r["Priority"] <= 3]

    ws2.merge_cells("A1:J1")
    t2 = ws2["A1"]
    t2.value = "🔥  TOP PICKS  —  Squeeze + Crossover Confluence"
    t2.font  = Font(name="Calibri", bold=True, size=13, color=YELLOW_SQZ)
    t2.fill  = cell_fill(DARK_BG)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    hdr2 = ["Symbol", "Signal", "Close ₹", "Spread%", "RSI(W)", "ADX(W)",
            "Vol Ratio", "Trend", "From52wH%", "TradingView"]
    cw2  = [14, 24, 11, 10, 9, 9, 10, 8, 12, 18]

    for col, (h, w) in enumerate(zip(hdr2, cw2), 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font      = hdr_font(size=9)
        c.fill      = cell_fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 24

    for i, r in enumerate(top_picks, 3):
        sig    = r["Signal"]
        bg, fg = SIGNAL_COLORS.get(sig, (HEADER_BG, WHITE))
        row_bg = "252538" if i % 2 == 0 else "1E1E2E"

        vals = [r["Symbol"], sig, r["Close"], r["Spread%"], r["RSI_W"],
                r["ADX_W"], r["Vol_Ratio"], r["Trend"],
                r["From52wHigh%"], f"NSE:{r['Symbol']}"]

        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font      = Font(name="Calibri", size=9, color=WHITE)
            c.fill      = cell_fill(row_bg)
            c.border    = thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws2.cell(row=i, column=2).fill = cell_fill(bg)
        ws2.cell(row=i, column=2).font = Font(name="Calibri", bold=True, size=9, color=fg)

        tv = ws2.cell(row=i, column=10)
        tv.hyperlink = f"https://www.tradingview.com/chart/?symbol=NSE:{r['Symbol']}"
        tv.font = Font(name="Calibri", size=9, color="4FC3F7", underline="single")

        trend_c = ws2.cell(row=i, column=8)
        trend_c.font = Font(name="Calibri", bold=True, size=9,
                            color=GREEN_BULL if r["Trend"]=="BULL" else "FF4444")

    ws2.freeze_panes = "A3"

    # ── Sheet 3: Legend ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("📖 Legend")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "4FC3F7"
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 55

    legend_data = [
        ("SIGNAL GUIDE", ""),
        ("🔥 SQUEEZE + CROSSOVER", "All 3 EMAs squeezing (<3% spread) + 10W crossing — STRONGEST signal"),
        ("⚡ SQUEEZE BREAKOUT",    "EMAs squeezed + 10W already above 30W & 40W — momentum building"),
        ("✅ DOUBLE CROSSOVER",    "10W crossed above both 30W and 40W in same week"),
        ("📈 10 X 30 CROSS",       "10W EMA crossed above 30W EMA this week"),
        ("📈 10 X 40 CROSS",       "10W EMA crossed above 40W EMA this week"),
        ("🔄 SQUEEZING",           "All 3 EMAs within 3% — watch for upcoming breakout"),
        ("", ""),
        ("COLUMN GUIDE", ""),
        ("Spread%",        "% gap between highest and lowest of 3 EMAs. Lower = tighter squeeze"),
        ("RSI(W)",         "Weekly RSI. >60 = strong momentum, 45-60 = building, <45 = weak"),
        ("ADX(W)",         "Weekly ADX. >25 = trending, <20 = ranging/accumulation"),
        ("Vol Ratio",      "Current volume / 20-week avg volume. >1.5 = strong confirmation"),
        ("Trend",          "BULL = Close above 200W EMA | BEAR = Close below 200W EMA"),
        ("From52wHigh%",   "How far below 52-week high. Closer to 0% = near highs"),
        ("", ""),
        ("BEST SETUP CRITERIA", ""),
        ("Signal",         "🔥 SQUEEZE + CROSSOVER (priority 1)"),
        ("RSI",            ">50, ideally >55 on weekly"),
        ("ADX",            ">20 and rising — confirms new trend starting"),
        ("Volume",         "Vol Ratio >1.5 on crossover week — institutional participation"),
        ("Trend",          "BULL preferred — price above 200W EMA"),
        ("Spread%",        "<2% is ideal — very tight squeeze before explosion"),
    ]

    ws3.row_dimensions[1].height = 22
    for i, (k, v) in enumerate(legend_data, 2):
        ck = ws3.cell(row=i, column=1, value=k)
        cv = ws3.cell(row=i, column=2, value=v)
        if k in ("SIGNAL GUIDE", "COLUMN GUIDE", "BEST SETUP CRITERIA"):
            ck.font  = Font(name="Calibri", bold=True, size=10, color=FIRE_ORANGE)
            ck.fill  = cell_fill(HEADER_BG)
            cv.fill  = cell_fill(HEADER_BG)
        elif k:
            bg_, fg_ = SIGNAL_COLORS.get(k, (DARK_BG, WHITE))
            ck.font  = Font(name="Calibri", size=9, color=fg_)
            ck.fill  = cell_fill(bg_)
            cv.font  = Font(name="Calibri", size=9, color=GREY_TEXT)
            cv.fill  = cell_fill("252538")
        for c in [ck, cv]:
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border    = thin_border()
        ws3.row_dimensions[i].height = 20

    wb.save(output_path)
    print(f"✅ Report saved: {output_path}")

# ── Entry Point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"🔍 Scanning {len(NIFTY500_STOCKS)} NSE stocks on Weekly timeframe...")
    print(f"   Signal: 10W EMA / 30W EMA / 40W EMA Squeeze + Crossover\n")

    results = []
    errors  = 0
    for i, sym in enumerate(NIFTY500_STOCKS, 1):
        r = screen_stock(sym)
        status = f"  [{i:>3}/{len(NIFTY500_STOCKS)}] {sym:<18}"
        if r:
            results.append(r)
            print(f"{status} ✅ {r['Signal']}")
        else:
            if i % 50 == 0:
                print(f"{status} —")

    print(f"\n{'─'*55}")
    print(f"📊 Stocks scanned : {len(NIFTY500_STOCKS)}")
    print(f"✅ Signals found  : {len(results)}")

    if results:
        # Summary by signal type
        from collections import Counter
        counts = Counter(r["Signal"] for r in results)
        print(f"\n📋 Signal Breakdown:")
        for sig, cnt in sorted(counts.items(), key=lambda x: x[0]):
            print(f"   {sig:<30} {cnt:>3} stocks")

        out = f"/mnt/user-data/outputs/NSE_EMA_Squeeze_{datetime.now().strftime('%d%b%Y')}.xlsx"
        os.makedirs(os.path.dirname(out), exist_ok=True)
        write_excel(results, out)
    else:
        print("⚠️  No signals found. Market may be trending without squeezes.")
