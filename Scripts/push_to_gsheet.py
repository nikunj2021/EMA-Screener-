"""
push_to_gsheet.py
─────────────────
1. Reads the latest row from each EMA history Excel file (Nifty50 / 250 / 500)
2. Appends new rows to the matching Google Sheets tab (idempotent — no duplicates)
3. Sends a single consolidated Telegram alert with all three breadth readings

Telegram message format:
────────────────────────
📊 *EMA Breadth Report — 09-Jun-2026*

🔹 *Nifty 50*
  %>20 EMA : 72.00%
  %>50 EMA : 65.00%
  %>200 EMA: 58.00%
  Total    : 50

🔹 *Nifty 250*  ...
🔹 *Nifty 500*  ...

🕐 Scan Time: 09-Jun-2026  19:00:12
"""

import os
import json
import gspread
import openpyxl
import requests
from google.oauth2.service_account import Credentials
from datetime import datetime, date, timedelta

# ── ENVIRONMENT VARIABLES (set as GitHub Secrets) ────────────────────────────
SHEET_ID          = os.environ["GSHEET_ID"]
CREDS_JSON        = os.environ["GSHEET_CREDS_JSON"]
TELEGRAM_BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_CHAT_ID   = os.environ["TELEGRAM_CHAT_ID"]

# ── EXCEL HISTORY FILES ───────────────────────────────────────────────────────
HISTORY_FILES = {
    "Nifty50_EMA":  {
        "path":     "Report/EMAReport/nifty_50_ema_breadth_history.xlsx",
        "label":    "Nifty 50",
        "sheet":    "EMA Breadth History",
    },
    "Nifty250_EMA": {
        "path":     "Report/EMAReport/Micro_250_ema_breadth_history.xlsx",
        "label":    "Nifty 250",
        "sheet":    "EMA Breadth History",
    },
    "Nifty500_EMA": {
        "path":     "Report/EMAReport/nifty_500_ema_breadth_history.xlsx",
        "label":    "Nifty 500",
        "sheet":    "EMA Breadth History",
    },
}

GSHEET_HEADERS = ["Date", "Scan Time", "Above EMA20 %", "Above EMA50 %", "Above EMA200 %", "Total Stocks"]

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ── GOOGLE SHEETS AUTH ────────────────────────────────────────────────────────
creds_dict = json.loads(CREDS_JSON)
creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
gc = gspread.authorize(creds)
sh = gc.open_by_key(SHEET_ID)


# ── HELPERS ───────────────────────────────────────────────────────────────────
def excel_date_to_str(value):
    """Convert Excel serial number OR date/datetime → DD-Mon-YYYY string."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(value))).strftime("%d-%b-%Y")
        except Exception:
            return str(value)
    return str(value)


def get_or_create_worksheet(sh, title, headers):
    """Return existing worksheet or create it with header row."""
    try:
        ws = sh.worksheet(title)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=1000, cols=20)
        ws.append_row(headers, value_input_option="USER_ENTERED")
        print(f"  Created new tab: {title}")
    return ws


def read_last_row(excel_path, sheet_name):
    """
    Read the last data row from an EMA history Excel file.
    Returns dict with keys: date, scan_time, pct20, pct50, pct200, total
    Columns by position: 0=Date, 1=ScanTime, 2=pct20, 3=pct50, 4=pct200, 5=total
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    print(f"  Sheets in file: {wb.sheetnames}")

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    print(f"  Total rows (incl. header): {len(rows)}")

    if len(rows) < 2:
        print(f"  ⚠ No data rows found.")
        return None

    last = rows[-1]
    print(f"  Last row raw: {last}")

    # Skip completely empty rows
    if all(v is None for v in last):
        print(f"  ⚠ Last row is empty — trying second-to-last row.")
        last = rows[-2]

    date_val  = excel_date_to_str(last[0])
    scan_time = str(last[1]).strip() if last[1] is not None else ""
    pct20     = round(float(last[2]), 2) if last[2] is not None else 0.0
    pct50     = round(float(last[3]), 2) if last[3] is not None else 0.0
    pct200    = round(float(last[4]), 2) if last[4] is not None else 0.0
    total     = int(last[5])             if last[5] is not None else 0

    return {
        "date":      date_val,
        "scan_time": scan_time,
        "pct20":     pct20,
        "pct50":     pct50,
        "pct200":    pct200,
        "total":     total,
    }


def emoji_for(pct):
    """Return colour emoji based on breadth percentage."""
    if pct >= 60:
        return "🟢"
    elif pct >= 40:
        return "🟡"
    else:
        return "🔴"


def send_telegram(message: str):
    """Send a Markdown-formatted message to Telegram."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id":    TELEGRAM_CHAT_ID,
        "text":       message,
        "parse_mode": "Markdown",
    }
    resp = requests.post(url, json=payload, timeout=15)
    if resp.status_code == 200:
        print("  ✅ Telegram alert sent.")
    else:
        print(f"  ❌ Telegram error {resp.status_code}: {resp.text}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
results = {}   # tab_name → row_dict  (used to build Telegram message)

for tab_name, cfg in HISTORY_FILES.items():
    print(f"\n{'─'*50}")
    print(f"Processing: {tab_name}  ({cfg['path']})")

    if not os.path.exists(cfg["path"]):
        print(f"  ⚠ File not found — skipping.")
        continue

    row = read_last_row(cfg["path"], cfg["sheet"])
    if not row:
        continue

    # ── Push to Google Sheets ─────────────────────────────────────────────────
    ws = get_or_create_worksheet(sh, tab_name, GSHEET_HEADERS)

    existing_dates = ws.col_values(1)   # all values in column A
    if row["date"] in existing_dates:
        print(f"  ℹ Already in sheet: {row['date']} — skipping duplicate.")
    else:
        new_row = [
            row["date"],
            row["scan_time"],
            row["pct20"],
            row["pct50"],
            row["pct200"],
            row["total"],
        ]
        ws.append_row(new_row, value_input_option="USER_ENTERED")
        print(f"  ✓ Appended to Google Sheets: {new_row}")

    results[tab_name] = row   # store for Telegram message

# ── BUILD & SEND TELEGRAM ALERT ───────────────────────────────────────────────
if results:
    # Use date from first available result
    first = next(iter(results.values()))
    report_date = first["date"]
    scan_time   = first["scan_time"]

    lines = [f"📊 *EMA Breadth Report — {report_date}*\n"]

    for tab_name, row in results.items():
        label = HISTORY_FILES[tab_name]["label"]
        lines.append(f"🔹 *{label}*")
        lines.append(f"  {emoji_for(row['pct20'])}  %>20 EMA  : `{row['pct20']:.2f}%`")
        lines.append(f"  {emoji_for(row['pct50'])}  %>50 EMA  : `{row['pct50']:.2f}%`")
        lines.append(f"  {emoji_for(row['pct200'])}  %>200 EMA : `{row['pct200']:.2f}%`")
        lines.append(f"  📌 Total Stocks : {row['total']}\n")

    lines.append(f"🕐 Scan Time: `{scan_time}`")

    message = "\n".join(lines)

    print(f"\n{'─'*50}")
    print("Sending Telegram alert...")
    print(message)
    send_telegram(message)

print(f"\n{'='*50}")
print("✅ push_to_gsheet.py complete.")