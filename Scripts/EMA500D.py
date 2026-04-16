"""
NSE EMA Breadth Analyzer — Nifty 50
Reads stock symbols from CSV, fetches daily closing prices & EMA data via yfinance,
calculates % of stocks trading above 20/50/200-day EMA, and saves results to Excel.

DATE SELECTION:
  On every run the script prompts:
    [T]  Today / latest available data
    [D]  Specific historical date
         Accepted formats: DD-Mon-YYYY | DD/MM/YYYY | YYYY-MM-DD
"""

import sys
import pandas as pd
import yfinance as yf
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import time
import os

warnings.filterwarnings("ignore")

# ─── FILE PATHS ──────────────────────────────────────────────────────────────
INPUT_CSV      = "StockList/nse500list.csv"
OUTPUT_EXCEL   = "Report/EMAReport/nifty_500_ema_breadth.xlsx"
HISTORY_EXCEL  = "Report/EMAReport/nifty_500_ema_breadth_history.xlsx"

# ─── SETTINGS ────────────────────────────────────────────────────────────────
BATCH_SIZE     = 20      # stocks per yfinance batch download
SLEEP_BETWEEN  = 1.5     # seconds between batches (avoid rate-limit)
MIN_BARS       = 210     # minimum history bars needed for 200-EMA


# ─── FILE SAVE HELPER ────────────────────────────────────────────────────────
def safe_save_xlsx(wb_or_df, path: str, is_df: bool = False) -> str:
    """
    Save an openpyxl Workbook (or DataFrame) to `path`.
    If the file is locked (open in Excel), automatically saves to a timestamped
    fallback filename in the same folder instead of crashing.
    Returns the actual path written.
    """
    def _write(target: str):
        if is_df:
            wb_or_df.to_excel(target, index=False)
        else:
            wb_or_df.save(target)

    try:
        _write(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        stamp     = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback  = f"{base}_{stamp}{ext}"
        print(f"\n  \u26a0  '{os.path.basename(path)}' is open in Excel \u2014 cannot overwrite.")
        print(f"     Saving to: {fallback}")
        _write(fallback)
        return fallback

# ─── RECENTLY LISTED STOCKS (< 1 year history) ───────────────────────────────
# Stocks listed < 1 year ago have sparse data with period="1y".
# Add newly-listed NSE stocks here in "SYMBOL.NS" format to fetch with max period.
RECENTLY_LISTED = {
    "EMCURE.NS",    # Listed Jul 2024
}


# ─── DATE PARSING ────────────────────────────────────────────────────────────
def parse_date_arg(raw: str) -> date:
    """
    Parse a user-supplied date string into a date object.
    Accepts:
      DD-Mon-YYYY  e.g.  12-Mar-2026
      DD-MM-YYYY   e.g.  12-03-2026
      DD/MM/YYYY   e.g.  12/03/2026
      YYYY-MM-DD   e.g.  2026-03-12
    """
    for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"Cannot parse date '{raw}'.\n"
        "  Accepted formats:\n"
        "    DD-Mon-YYYY  →  12-Mar-2026\n"
        "    DD-MM-YYYY   →  12-03-2026\n"
        "    DD/MM/YYYY   →  12/03/2026\n"
        "    YYYY-MM-DD   →  2026-03-12"
    )


def prompt_date() -> date | None:
    """
    Interactively ask the user whether they want today's data or a specific date.
    Returns a date object (for historical) or None (for today/latest).
    """
    print("\n  ┌─────────────────────────────────────────────┐")
    print("  │           DATE SELECTION                    │")
    print("  ├─────────────────────────────────────────────┤")
    print("  │  [T]  Today / latest available data         │")
    print("  │  [D]  Specific historical date              │")
    print("  └─────────────────────────────────────────────┘")
    choice = input("\n  Enter choice (T / D): ").strip().upper()

    if choice == "D":
        raw = input("  Enter date (DD-Mon-YYYY | DD-MM-YYYY | DD/MM/YYYY | YYYY-MM-DD): ").strip()
        d = parse_date_arg(raw)
        if d > date.today():
            print(f"  ⚠  Warning: {d.strftime('%d-%b-%Y')} is a future date — data may not be available.")
        return d
    else:
        return None     # means "use latest"


def resolve_fetch_window(target_date: date | None) -> tuple[str, str]:
    """
    Given the target date, decide the yfinance start/end window to download.
    We fetch ~465 calendar days before the target so that the 200-day EMA
    has sufficient warm-up history (~300 trading days).

    Returns (start_str, end_str) in YYYY-MM-DD format.
    yfinance end is EXCLUSIVE, so end = target_date + 1.
    """
    end_dt   = (target_date if target_date else date.today()) + timedelta(days=1)
    start_dt = end_dt - timedelta(days=365 + 100)   # ~465 calendar days
    return start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d")


# ─── HELPERS ─────────────────────────────────────────────────────────────────
def load_symbols(csv_path: str) -> list[str]:
    """Load symbols from CSV. Accepts common column names."""
    df = pd.read_csv(csv_path)
    df.columns = df.columns.str.strip()
    for col in ["Symbol", "symbol", "SYMBOL", "Ticker", "ticker", "TICKER", "NSE Symbol"]:
        if col in df.columns:
            syms = df[col].dropna().str.strip().tolist()
            return [s + ".NS" if not s.endswith(".NS") else s for s in syms]
    raise ValueError(f"No symbol column found. Available columns: {df.columns.tolist()}")


def slice_to_date(series: pd.Series, target_date: date | None) -> pd.Series:
    """
    If target_date is set, keep only rows on or before that date.
    Handles both timezone-aware and naive DatetimeIndex.
    Raises ValueError if no data remains after slicing.
    """
    if target_date is None or series.empty:
        return series

    # Convert index dates to plain date objects for comparison
    idx_dates = [
        d.date() if hasattr(d, "date") else d
        for d in series.index
    ]
    mask = [d <= target_date for d in idx_dates]
    sliced = series[mask]

    if sliced.empty:
        raise ValueError(f"No data on or before {target_date.strftime('%d-%b-%Y')}")
    return sliced


def fetch_single(sym: str, start: str, end: str) -> pd.Series | None:
    """Fetch close prices for one symbol using start/end window, with fallbacks."""
    # Method 1: yf.download with date range
    try:
        raw = yf.download(sym, start=start, end=end, interval="1d",
                          auto_adjust=True, progress=False, repair=False)
        if raw is not None and not raw.empty:
            col = raw["Close"] if "Close" in raw.columns else raw.iloc[:, 0]
            s = col.dropna()
            if len(s) > 0:
                return s
    except Exception:
        pass

    # Method 2: yf.Ticker with broader period fallback
    for p in ["2y", "max"]:
        try:
            t = yf.Ticker(sym)
            hist = t.history(period=p, interval="1d", auto_adjust=True)
            if hist is not None and not hist.empty and "Close" in hist.columns:
                s = hist["Close"].dropna()
                if len(s) > 0:
                    return s
        except Exception:
            pass

    return None


def fetch_in_batches(symbols: list[str], start: str, end: str) -> dict:
    """Download daily OHLCV for all symbols in batches, with per-ticker retry."""
    all_closes = {}
    failed_batch = []

    # ── Pre-fetch recently-listed stocks individually ─────────────────────────
    recent = [s for s in symbols if s in RECENTLY_LISTED]
    normal = [s for s in symbols if s not in RECENTLY_LISTED]
    if recent:
        print(f"\n  Pre-fetching {len(recent)} recently-listed stock(s)...")
        for sym in recent:
            print(f"    → {sym}", end="  ")
            s = fetch_single(sym, start, end)
            if s is not None:
                all_closes[sym] = s
                print(f"OK  ({len(s)} bars)")
            else:
                print("FAILED — will skip")
        time.sleep(0.5)

    total_batches = -(-len(normal) // BATCH_SIZE)

    for i in range(0, len(normal), BATCH_SIZE):
        batch = normal[i : i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Fetching batch {batch_num}/{total_batches}  ({len(batch)} symbols)...")
        try:
            raw = yf.download(
                batch,
                start=start,
                end=end,
                interval="1d",
                group_by="ticker",
                auto_adjust=True,
                progress=False,
                threads=True,
                repair=False,
            )
            if raw is None or raw.empty:
                failed_batch.extend(batch)
            elif len(batch) == 1:
                sym = batch[0]
                try:
                    col = raw["Close"] if "Close" in raw.columns else raw.iloc[:, 0]
                    s = col.dropna()
                    if len(s) > 0:
                        all_closes[sym] = s
                    else:
                        failed_batch.append(sym)
                except Exception:
                    failed_batch.append(sym)
            else:
                # Multi-level columns: (field, ticker)
                for sym in batch:
                    try:
                        if isinstance(raw.columns, pd.MultiIndex):
                            if ("Close", sym) in raw.columns:
                                s = raw[("Close", sym)].dropna()
                            elif sym in raw.columns.get_level_values(1):
                                s = raw.xs(sym, axis=1, level=1)["Close"].dropna()
                            else:
                                failed_batch.append(sym)
                                continue
                        else:
                            s = raw[sym]["Close"].dropna() if sym in raw else pd.Series()

                        if len(s) > 0:
                            all_closes[sym] = s
                        else:
                            failed_batch.append(sym)
                    except Exception:
                        failed_batch.append(sym)
        except Exception as e:
            print(f"    Batch error: {e}")
            failed_batch.extend(batch)

        time.sleep(SLEEP_BETWEEN)

    # ── Retry failed tickers one-by-one ──────────────────────────────────────
    if failed_batch:
        print(f"\n  Retrying {len(failed_batch)} failed symbol(s) individually...")
        still_failed = []
        for sym in failed_batch:
            print(f"    → {sym}", end="  ")
            s = fetch_single(sym, start, end)
            if s is not None:
                all_closes[sym] = s
                print("OK")
            else:
                still_failed.append(sym)
                print("FAILED")
            time.sleep(0.5)
        if still_failed:
            print(f"\n  WARNING: {len(still_failed)} symbol(s) could not be fetched and will be skipped:")
            print(f"  {still_failed}")

    return all_closes


def compute_ema(series: pd.Series, span: int) -> float | None:
    """
    Return the latest EMA value of the (already-sliced) series.
    Returns None if data is insufficient.
    """
    min_required = max(span // 2, 20)
    if len(series) < min_required:
        return None
    val = series.ewm(span=span, adjust=False).mean().iloc[-1]
    # Ensure scalar — newer pandas/yfinance may return a Series
    if isinstance(val, pd.Series):
        val = val.iloc[0]
    return float(val)


def analyse(closes_dict: dict, target_date: date | None) -> list[dict]:
    """
    For each symbol:
      1. Slice the close series to target_date (if specified)
      2. Use the last available closing price on or before target_date
      3. Compute EMAs on the sliced history
    """
    results = []
    skipped = 0

    for sym, closes in closes_dict.items():
        try:
            sliced = slice_to_date(closes, target_date)
        except ValueError:
            skipped += 1
            continue

        # Extract scalar closing price
        price = sliced.iloc[-1]
        if isinstance(price, pd.Series):
            price = price.iloc[0]
        price = float(price)

        # Previous day close (second-to-last bar, if available)
        if len(sliced) >= 2:
            prev = sliced.iloc[-2]
            if isinstance(prev, pd.Series):
                prev = prev.iloc[0]
            prev_close = float(prev)
        else:
            prev_close = None

        # Record the actual date used (last trading day on/before target)
        raw_date = sliced.index[-1]
        actual_date = raw_date.date() if hasattr(raw_date, "date") else raw_date

        e20  = compute_ema(sliced, 20)
        e50  = compute_ema(sliced, 50)
        e200 = compute_ema(sliced, 200)

        results.append({
            "Symbol":        sym,
            "Price Date":    actual_date.strftime("%d-%b-%Y"),
            "Closing Price": round(price, 2),
            "Prev Close":    round(prev_close, 2) if prev_close is not None else None,
            "EMA20":         round(e20,  2) if e20  is not None else None,
            "EMA50":         round(e50,  2) if e50  is not None else None,
            "EMA200":        round(e200, 2) if e200 is not None else None,
            "Above EMA20":   bool(price > e20)  if e20  is not None else False,
            "Above EMA50":   bool(price > e50)  if e50  is not None else False,
            "Above EMA200":  bool(price > e200) if e200 is not None else False,
        })

    if skipped:
        print(f"  ⚠  {skipped} symbol(s) had no data on/before the target date — skipped.")

    return results


def pct(count, total: int) -> float:
    return round(count / total * 100, 2) if total else 0.0


# ─── EXCEL STYLING ───────────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", start_color="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF", size=11, name="Arial")
BOLD        = Font(bold=True, name="Arial", size=11)
NORMAL      = Font(name="Arial", size=10)
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
RED_FILL    = PatternFill("solid", start_color="FFC7CE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")
BLUE_FILL   = PatternFill("solid", start_color="D9E2F3")
THIN_SIDE   = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")


def style_header_row(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.border    = THIN_BORDER
        cell.alignment = CENTER


def write_summary_sheet(wb: openpyxl.Workbook, scan_time: str, report_date_label: str,
                         breadth: dict, detail_df: pd.DataFrame):
    ws = wb.active
    ws.title = "EMA Breadth Summary"

    # ── Title ──
    ws.merge_cells("A1:F1")
    ws["A1"] = "NSE EMA MARKET BREADTH ANALYSIS"
    ws["A1"].font      = Font(bold=True, size=14, color="1F4E79", name="Arial")
    ws["A1"].alignment = CENTER
    ws["A1"].fill      = BLUE_FILL

    # ── Sub-title: scan run time + report date ──
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Scan Run: {scan_time}    |    Closing Price Date: {report_date_label}"
    ws["A2"].font      = Font(italic=True, size=10, name="Arial", color="595959")
    ws["A2"].alignment = CENTER

    ws.append([])  # blank row 3

    # ── Breadth Table Header ──
    headers = ["Closing Price Date", "% Above 20 EMA", "% Above 50 EMA",
               "% Above 200 EMA", "Total Stocks Analysed", "Scan Run At"]
    ws.append(headers)
    style_header_row(ws, 4, len(headers))

    # ── Breadth Data Row ──
    ws.append([
        report_date_label,
        breadth["pct20"],
        breadth["pct50"],
        breadth["pct200"],
        breadth["total"],
        scan_time,
    ])
    data_row = 5
    for c in range(1, 7):
        cell = ws.cell(row=data_row, column=c)
        cell.font      = BOLD
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
        if c in (2, 3, 4):
            cell.number_format = "0.00\"%\""
            val = cell.value
            if val is not None:
                if val >= 60:
                    cell.fill = GREEN_FILL
                elif val >= 40:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL

    ws.append([])  # blank

    # ── Market Breadth Interpretation ──
    ws.append(["Market Breadth Interpretation"])
    ws.cell(row=7, column=1).font = Font(bold=True, size=11, name="Arial", color="1F4E79")

    def interpret(pct_val, ema_label):
        if pct_val >= 70:
            return f"{ema_label}: STRONG BULLISH — {pct_val}% stocks above EMA (market broadly trending up)"
        elif pct_val >= 55:
            return f"{ema_label}: MODERATELY BULLISH — {pct_val}% stocks above EMA"
        elif pct_val >= 45:
            return f"{ema_label}: NEUTRAL — {pct_val}% stocks above EMA (mixed market)"
        elif pct_val >= 30:
            return f"{ema_label}: MODERATELY BEARISH — {pct_val}% stocks above EMA"
        else:
            return f"{ema_label}: STRONGLY BEARISH — {pct_val}% stocks above EMA (market broadly declining)"

    for line in [
        interpret(breadth["pct20"],  "Short-Term  (20 EMA)"),
        interpret(breadth["pct50"],  "Medium-Term (50 EMA)"),
        interpret(breadth["pct200"], "Long-Term  (200 EMA)"),
    ]:
        ws.append([line])
        ws.cell(row=ws.max_row, column=1).font      = NORMAL
        ws.cell(row=ws.max_row, column=1).alignment = LEFT

    ws.append([])

    # ── Counts ──
    ws.append(["Counts"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial", color="1F4E79")
    ws.append([f"  Above 20 EMA : {breadth['cnt20']} of {breadth['total']} stocks"])
    ws.append([f"  Above 50 EMA : {breadth['cnt50']} of {breadth['total']} stocks"])
    ws.append([f"  Above 200 EMA: {breadth['cnt200']} of {breadth['total']} stocks"])
    for r in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = NORMAL

    # ── Column widths ──
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 24
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A5"


def write_detail_sheet(wb: openpyxl.Workbook, detail_df: pd.DataFrame):
    ws = wb.create_sheet("Stock Detail")

    cols = ["Symbol", "Price Date", "Closing Price", "EMA20", "EMA50", "EMA200",
            "Above EMA20", "Above EMA50", "Above EMA200", "TradingView Chart"]
    ws.append(cols)
    style_header_row(ws, 1, len(cols))

    TV_FONT       = Font(name="Arial", size=10, color="1155CC", underline="single")
    TV_FONT_BOLD  = Font(name="Arial", size=10, color="1155CC", underline="single", bold=False)

    for _, row in detail_df.iterrows():
        clean_sym = row["Symbol"].replace(".NS", "")
        tv_url    = f"https://www.tradingview.com/chart/?symbol=NSE:{clean_sym}"

        ws.append([
            clean_sym,
            row["Price Date"],
            row["Closing Price"],
            row["EMA20"],
            row["EMA50"],
            row["EMA200"],
            "YES" if row["Above EMA20"]  else "NO",
            "YES" if row["Above EMA50"]  else "NO",
            "YES" if row["Above EMA200"] else "NO",
            clean_sym,          # display text for hyperlink cell
        ])
        r = ws.max_row
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font      = NORMAL
            cell.border    = THIN_BORDER
            cell.alignment = CENTER if c > 1 else LEFT
        for c_idx, col_name in [(7, "Above EMA20"), (8, "Above EMA50"), (9, "Above EMA200")]:
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = GREEN_FILL if row[col_name] else RED_FILL

        # ── TradingView hyperlink in column 10 ──
        tv_cell            = ws.cell(row=r, column=10)
        tv_cell.hyperlink  = tv_url
        tv_cell.value      = clean_sym
        tv_cell.font       = TV_FONT
        tv_cell.alignment  = CENTER

    widths = [14, 14, 16, 12, 12, 12, 14, 14, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


SHEET_EMA_HIST = "EMA Breadth History"
_EMA_HIST_COLS = ["Date", "Scan Run At", "% >20 EMA", "% >50 EMA", "% >200 EMA", "Total Stocks"]

# ─── PRICE-CHANGE BREADTH HISTORY ───────────────────────────────────────────
SHEET_BREADTH = "nifty 50_ema_breadth_history"

# Column headers for the price-change breadth sheet
_BREADTH_COLS = [
    "Date",
    "Dn 0 to -2.99%",
    "Dn -3 to -4.99%",
    "Dn ≥-5%",
    "Flat (0%)",
    "Up 0 to 2.99%",
    "Up 3 to 4.99%",
    "Up ≥5%",
]

# Colour scheme for each bucket column (indices 1-7, 0-based within data cols)
_BUCKET_COLOURS = [
    "FFC7CE",   # light red   — down 0 to -2.99%
    "FF0000",   # red         — down -3 to -4.99%
    "C00000",   # dark red    — down ≥ -5%
    "D9D9D9",   # grey        — flat
    "C6EFCE",   # light green — up 0 to 2.99%
    "00B050",   # green       — up 3 to 4.99%
    "375623",   # dark green  — up ≥ 5%
]


def _compute_breadth_buckets(detail_df: pd.DataFrame) -> dict:
    """
    Given the detail DataFrame (must contain 'Closing Price' and 'Prev Close'),
    return bucket percentages rounded to 2 dp.
    Stocks with no prev-close data are excluded from totals.
    """
    valid = detail_df.dropna(subset=["Prev Close"]).copy()
    valid = valid[valid["Prev Close"] > 0]        # guard against div/0
    n = len(valid)
    if n == 0:
        return {k: 0.0 for k in _BREADTH_COLS[1:]}

    chg = (valid["Closing Price"] - valid["Prev Close"]) / valid["Prev Close"] * 100

    def _pct(mask):
        return round(mask.sum() / n * 100, 2)

    return {
        _BREADTH_COLS[1]: _pct((chg < 0)     & (chg >= -2.99)),   # -2.99% to < 0%
        _BREADTH_COLS[2]: _pct((chg < -2.99) & (chg >= -4.99)),   # -4.99% to < -2.99%
        _BREADTH_COLS[3]: _pct(chg <= -5.0),                       # ≤ -5%
        _BREADTH_COLS[4]: _pct(chg == 0.0),                        # exactly 0%
        _BREADTH_COLS[5]: _pct((chg > 0)     & (chg < 3.00)),      # > 0% to < 3%
        _BREADTH_COLS[6]: _pct((chg >= 3.00) & (chg < 5.00)),      # 3% to < 5%
        _BREADTH_COLS[7]: _pct(chg >= 5.0),                        # ≥ 5%
    }


def _write_or_update_ema_sheet(wb, report_date_label: str, scan_time: str, breadth: dict):
    """Write/append one row to the EMA Breadth History sheet."""
    if SHEET_EMA_HIST in wb.sheetnames:
        ws = wb[SHEET_EMA_HIST]
    else:
        ws = wb.create_sheet(SHEET_EMA_HIST, 0)
        ws.append(_EMA_HIST_COLS)
        style_header_row(ws, 1, len(_EMA_HIST_COLS))
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 24
        for col_idx in range(3, len(_EMA_HIST_COLS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_EMA_HIST_COLS))}1"

    ws.append([report_date_label, scan_time,
               breadth["pct20"], breadth["pct50"], breadth["pct200"], breadth["total"]])
    r = ws.max_row
    for c in range(1, len(_EMA_HIST_COLS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font      = NORMAL
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
        if c in (3, 4, 5):
            cell.number_format = "0.00\"%\""
            val = cell.value
            if val is not None:
                cell.fill = GREEN_FILL if val >= 60 else (YELLOW_FILL if val >= 40 else RED_FILL)


def _write_or_update_breadth_sheet(wb, report_date_label: str, detail_df: pd.DataFrame):
    """Write/append one row to the price-change breadth history sheet."""
    buckets = _compute_breadth_buckets(detail_df)
    new_row = {_BREADTH_COLS[0]: report_date_label}
    new_row.update(buckets)

    if SHEET_BREADTH in wb.sheetnames:
        ws = wb[SHEET_BREADTH]
    else:
        ws = wb.create_sheet(SHEET_BREADTH)
        ws.append(_BREADTH_COLS)
        style_header_row(ws, 1, len(_BREADTH_COLS))
        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(_BREADTH_COLS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_BREADTH_COLS))}1"

    data_row = [new_row[col] for col in _BREADTH_COLS]
    ws.append(data_row)
    r = ws.max_row

    ws.cell(row=r, column=1).font      = NORMAL
    ws.cell(row=r, column=1).border    = THIN_BORDER
    ws.cell(row=r, column=1).alignment = CENTER

    for col_idx, colour in enumerate(_BUCKET_COLOURS, start=2):
        cell = ws.cell(row=r, column=col_idx)
        cell.number_format = "0.00\"%\""
        cell.border        = THIN_BORDER
        cell.alignment     = CENTER
        cell.font          = Font(name="Arial", size=10,
                                  color="FFFFFF" if colour in ("FF0000", "C00000", "375623", "00B050") else "000000")
        cell.fill = PatternFill("solid", start_color=colour)


def append_history_both(history_path: str, scan_time: str, report_date_label: str,
                        breadth: dict, detail_df: pd.DataFrame):
    """
    Load (or create) the history workbook once, write both history sheets,
    and save once — so neither sheet ever overwrites the other.
    """
    if os.path.exists(history_path):
        wb = openpyxl.load_workbook(history_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _write_or_update_ema_sheet(wb, report_date_label, scan_time, breadth)
    _write_or_update_breadth_sheet(wb, report_date_label, detail_df)

    actual = safe_save_xlsx(wb, history_path)
    print(f"  History updated → {actual}")
    print(f"    Sheets: '{SHEET_EMA_HIST}' + '{SHEET_BREADTH}'")


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    scan_time = datetime.now().strftime("%d-%b-%Y  %H:%M:%S")

    # ── Always prompt interactively for date ─────────────────────────────────
    target_date: date | None = prompt_date()

    if target_date:
        report_date_label = target_date.strftime("%d-%b-%Y")
        mode_label        = f"Historical — closing price as of {report_date_label}"
    else:
        report_date_label = "Latest (Today)"
        mode_label        = "Today / Latest Available Closing Price"

    print(f"\n{'='*65}")
    print(f"  NSE EMA Breadth Analyser")
    print(f"  Scan Run  : {scan_time}")
    print(f"  Mode      : {mode_label}")
    print(f"{'='*65}")

    # ── Determine fetch window ────────────────────────────────────────────────
    start_str, end_str = resolve_fetch_window(target_date)
    print(f"\n  Data window : {start_str}  →  {end_str}")

    # 1. Load symbols
    print("\n[1] Loading symbols from CSV...")
    symbols = load_symbols(INPUT_CSV)
    print(f"    {len(symbols)} symbols loaded.")

    # 2. Fetch data
    print("\n[2] Fetching daily closing prices from Yahoo Finance...")
    closes_dict = fetch_in_batches(symbols, start_str, end_str)
    print(f"    Data received for {len(closes_dict)} symbols.")

    # 3. Analyse
    print("\n[3] Slicing to target date and computing EMAs...")
    results   = analyse(closes_dict, target_date)
    detail_df = pd.DataFrame(results)

    if detail_df.empty:
        print("\n  ERROR: No data available for the specified date. "
              "It may be a market holiday or weekend — try the nearest trading day.")
        sys.exit(1)

    # Determine the actual closing date used in the report
    # (may differ slightly across stocks due to holidays)
    unique_dates = detail_df["Price Date"].unique()
    if len(unique_dates) == 1:
        actual_report_date = unique_dates[0]
        print(f"    Closing price date : {actual_report_date}")
    else:
        # Use the most common date
        actual_report_date = detail_df["Price Date"].mode()[0]
        print(f"    Closing price dates (multiple due to holidays): {sorted(unique_dates)}")
        print(f"    Primary date used in report                  : {actual_report_date}")

    total  = len(detail_df)
    cnt20  = detail_df["Above EMA20"].sum()
    cnt50  = detail_df["Above EMA50"].sum()
    cnt200 = detail_df["Above EMA200"].sum()

    breadth = {
        "total":  total,
        "cnt20":  int(cnt20),
        "cnt50":  int(cnt50),
        "cnt200": int(cnt200),
        "pct20":  pct(cnt20,  total),
        "pct50":  pct(cnt50,  total),
        "pct200": pct(cnt200, total),
    }

    print(f"\n    {'Metric':<30} {'Value':>10}")
    print(f"    {'-'*42}")
    print(f"    {'Total stocks analysed':<30} {total:>10}")
    print(f"    {'Above 20-day EMA':<30} {cnt20:>5} ({breadth['pct20']:>5.1f}%)")
    print(f"    {'Above 50-day EMA':<30} {cnt50:>5} ({breadth['pct50']:>5.1f}%)")
    print(f"    {'Above 200-day EMA':<30} {cnt200:>5} ({breadth['pct200']:>5.1f}%)")

    # 4. Write Excel
    print(f"\n[4] Writing Excel report → {OUTPUT_EXCEL}")
    wb = openpyxl.Workbook()
    write_summary_sheet(wb, scan_time, actual_report_date, breadth, detail_df)
    write_detail_sheet(wb, detail_df)
    saved_path = safe_save_xlsx(wb, OUTPUT_EXCEL)
    print(f"    Done → {saved_path}")

    # 5. Append to history (both sheets in one workbook load/save)
    print("\n[5] Appending to history log...")
    append_history_both(HISTORY_EXCEL, scan_time, actual_report_date, breadth, detail_df)

    print(f"\n{'='*65}")
    print("  Analysis complete!")
    print(f"  Closing Price Date : {actual_report_date}")
    print(f"  Main report        : {saved_path}")
    print(f"  History log        : {HISTORY_EXCEL}")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
