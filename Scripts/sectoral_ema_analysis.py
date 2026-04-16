"""
Sectoral EMA Analysis for NSE Stocks
Analyzes stocks trading above 20, 50, 200-day EMA by sector
"""

import pandas as pd
import numpy as np
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
import time
from datetime import datetime

warnings.filterwarnings('ignore')

# ── CONFIG ──────────────────────────────────────────────────────────────────
CSV_PATH   = "StockList/SectoreStocks.csv"
OUTPUT_PATH = "Report/EMAReport/Sectoral_EMA_Analysis.xlsx"
PERIOD     = "1y"        # fetch 1 year of daily data

# ── STYLE HELPERS ────────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row, cols, text_list, fill_hex, font_color="FFFFFF",
                 font_size=11, bold=True):
    fill = hdr_fill(fill_hex)
    font = Font(bold=bold, color=font_color, size=font_size, name="Arial")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, text in zip(cols, text_list):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
        cell.border = thin_border()

def fmt_cell(cell, number_format=None, align_h="center", bold=False,
             font_color="000000", fill_hex=None, font_size=10):
    cell.font      = Font(name="Arial", size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align_h, vertical="center")
    cell.border    = thin_border()
    if number_format:
        cell.number_format = number_format
    if fill_hex:
        cell.fill = hdr_fill(fill_hex)

# ── STEP 1 : Load CSV ────────────────────────────────────────────────────────
print("Loading CSV …")
df_stocks = pd.read_csv(CSV_PATH)
df_stocks.columns = df_stocks.columns.str.strip()
df_stocks["Ticker"] = df_stocks["Ticker"].str.strip()
df_stocks["Sector"] = df_stocks["Sector"].str.strip()

tickers_nse = (df_stocks["Ticker"] + ".NS").tolist()
ticker_to_sector = dict(zip(df_stocks["Ticker"] + ".NS", df_stocks["Sector"]))
all_sectors = df_stocks["Sector"].unique().tolist()

print(f"  → {len(df_stocks)} stocks across {len(all_sectors)} sectors")

# ── STEP 2 : Fetch price data in batches ────────────────────────────────────
print(f"\nDownloading price data (period={PERIOD}) …")

price_data = {}   # ticker → closing price Series
failed     = []

for idx, ticker in enumerate(tickers_nse, 1):
    s = None

    # --- Method 1: yf.Ticker().history() — most reliable for NSE ---
    try:
        tk  = yf.Ticker(ticker)
        raw = tk.history(period=PERIOD, interval="1d", auto_adjust=True)
        if raw is not None and not raw.empty and "Close" in raw.columns:
            candidate = raw["Close"].dropna()
            if len(candidate) >= 20:
                s = candidate
    except Exception:
        pass

    # --- Method 2: yf.download() fallback ---
    if s is None:
        try:
            raw = yf.download(ticker, period=PERIOD, interval="1d",
                              auto_adjust=True, progress=False, threads=False)
            if raw is not None and not raw.empty:
                if isinstance(raw.columns, pd.MultiIndex):
                    close_col = ("Close", ticker)
                    if close_col in raw.columns:
                        candidate = raw[close_col].dropna()
                        if len(candidate) >= 20:
                            s = candidate
                elif "Close" in raw.columns:
                    candidate = raw["Close"].dropna()
                    if len(candidate) >= 20:
                        s = candidate
        except Exception:
            pass

    if s is not None:
        price_data[ticker] = s
    else:
        failed.append(ticker)

    if idx % 25 == 0 or idx == len(tickers_nse):
        print(f"  [{idx:>3}/{len(tickers_nse)}] fetched={len(price_data)}  failed={len(failed)}")

    time.sleep(0.3)

if failed:
    print(f"\n  ⚠  Could not fetch data for {len(failed)} ticker(s): {', '.join(t.replace('.NS','') for t in failed)}")
print(f"\n  → Data fetched for {len(price_data)} / {len(tickers_nse)} tickers")

# ── STEP 3 : Compute EMA flags ───────────────────────────────────────────────
def calc_ema(series, span):
    return series.ewm(span=span, adjust=False).mean()

records = []   # detailed per-ticker rows for Sheet 2

for ticker, closes in price_data.items():
    last_price = closes.iloc[-1]
    ema20  = calc_ema(closes, 20).iloc[-1]
    ema50  = calc_ema(closes, 50).iloc[-1] if len(closes) >= 50 else np.nan
    ema200 = calc_ema(closes, 200).iloc[-1] if len(closes) >= 200 else np.nan

    above20  = (last_price > ema20)  if not np.isnan(ema20)  else False
    above50  = (last_price > ema50)  if not np.isnan(ema50)  else False
    above200 = (last_price > ema200) if not np.isnan(ema200) else False

    sector = ticker_to_sector.get(ticker, "Unknown")
    symbol = ticker.replace(".NS", "")

    records.append({
        "Sector"  : sector,
        "Ticker"  : symbol,
        "Last"    : round(last_price, 2),
        "EMA20"   : round(ema20,  2) if not np.isnan(ema20)  else None,
        "EMA50"   : round(ema50,  2) if not np.isnan(ema50)  else None,
        "EMA200"  : round(ema200, 2) if not np.isnan(ema200) else None,
        "Above20" : above20,
        "Above50" : above50,
        "Above200": above200,
        "Above_All": above20 and above50 and above200,
    })

df_detail = pd.DataFrame(records)

# ── STEP 4 : Sector-level summary ────────────────────────────────────────────
# Total stocks per sector (from CSV, not just those with data)
total_map = df_stocks.groupby("Sector")["Ticker"].count().to_dict()

rows_summary = []
for sector in all_sectors:
    sub   = df_detail[df_detail["Sector"] == sector]
    total = total_map.get(sector, 0)
    n20   = int(sub["Above20"].sum())
    n50   = int(sub["Above50"].sum())
    n200  = int(sub["Above200"].sum())
    pct20  = round(n20  / total * 100, 1) if total else 0
    pct50  = round(n50  / total * 100, 1) if total else 0
    pct200 = round(n200 / total * 100, 1) if total else 0
    rows_summary.append({
        "Sector"            : sector,
        "Total_Stocks"      : total,
        "Above_20EMA"       : n20,
        "Above_50EMA"       : n50,
        "Above_200EMA"      : n200,
        "Pct_Above_20EMA"   : pct20,
        "Pct_Above_50EMA"   : pct50,
        "Pct_Above_200EMA"  : pct200,
        # Composite strength = average of three percentages
        "Strength"         : round((pct20 + pct50 + pct200) / 3, 2),
    })

df_summary = pd.DataFrame(rows_summary).sort_values("Strength", ascending=False)

# ── STEP 5 : Build Excel workbook ────────────────────────────────────────────
print("\nBuilding Excel report …")
wb = Workbook()

# ─── colour palette ───────────────────────────────────────────────────────────
C_TITLE_BG   = "1F3864"   # dark navy  – title rows
C_HDR_BG     = "2E75B6"   # mid blue   – column headers
C_HDR_GRP    = "4472C4"   # lighter blue – group sub-headers
C_ALT        = "EBF3FB"   # very light blue – alternating data rows
C_SEC_BAND   = "D6E4F7"   # sector-band highlight (sheet 2)
C_WHITE      = "FFFFFF"
C_TEXT       = "1A1A1A"

def title_style(cell, text, span_end_col):
    cell.value     = text
    cell.font      = Font(name="Arial", size=12, bold=True, color=C_WHITE)
    cell.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def hdr_style(cell, text, bg=C_HDR_BG):
    cell.value     = text
    cell.font      = Font(name="Arial", size=10, bold=True, color=C_WHITE)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        left=Side(style="thin", color="AAAAAA"), right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),  bottom=Side(style="medium", color="555555"))

def data_style(cell, val, fmt=None, align="center", bold=False, bg=C_WHITE):
    cell.value     = val
    cell.font      = Font(name="Arial", size=10, bold=bold, color=C_TEXT)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"))
    if fmt:
        cell.number_format = fmt

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Sector EMA Summary
# Cols: Sector | Total Stocks | #>20EMA | #>50EMA | #>200EMA | %>20EMA | %>50EMA | %>200EMA | Composite%
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sector EMA Summary"
ws1.sheet_view.showGridLines = False

# ── row heights ──────────────────────────────────────────────────────────────
ws1.row_dimensions[1].height = 28   # main title
ws1.row_dimensions[2].height = 18   # group labels
ws1.row_dimensions[3].height = 36   # column headers

# ── column widths ─────────────────────────────────────────────────────────────
S1_COLS = [("A", 30), ("B", 10), ("C", 12), ("D", 12), ("E", 12),
           ("F", 12),  ("G", 12), ("H", 12), ("I", 14)]
for col, w in S1_COLS:
    ws1.column_dimensions[col].width = w

# ── Row 1 : Main title ────────────────────────────────────────────────────────
ws1.merge_cells("A1:I1")
title_style(ws1["A1"],
    f"NSE Sectoral EMA Analysis   ·   Generated: {datetime.now().strftime('%d-%b-%Y  %H:%M')}",
    9)

# ── Row 2 : Group labels ──────────────────────────────────────────────────────
ws1.merge_cells("C2:E2")
ws1.merge_cells("F2:H2")
for cell_ref, label in [("C2", "No. of Stocks Above EMA"), ("F2", "% of Stocks Above EMA")]:
    c = ws1[cell_ref]
    c.value     = label
    c.font      = Font(name="Arial", size=9, bold=True, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=C_HDR_GRP)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = Border(bottom=Side(style="thin", color="AAAAAA"))

# ── Row 3 : Column headers ────────────────────────────────────────────────────
HDR1 = ["Sector", "Total\nStocks",
        "Above\n20 EMA", "Above\n50 EMA", "Above\n200 EMA",
        "% Above\n20 EMA", "% Above\n50 EMA", "% Above\n200 EMA",
        "Composite\nStrength %"]
for ci, h in enumerate(HDR1, 1):
    hdr_style(ws1.cell(row=3, column=ci), h)

# ── Rows 4+ : Data ────────────────────────────────────────────────────────────
for ri, row in enumerate(df_summary.itertuples(index=False), start=4):
    bg = C_ALT if ri % 2 == 0 else C_WHITE
    pct20  = row.Pct_Above_20EMA  / 100
    pct50  = row.Pct_Above_50EMA  / 100
    pct200 = row.Pct_Above_200EMA / 100
    comp   = row.Strength         / 100

    row_data = [
        (row.Sector,       None,    "left",   True),
        (row.Total_Stocks, "0",     "center", False),
        (row.Above_20EMA,  "0",     "center", False),
        (row.Above_50EMA,  "0",     "center", False),
        (row.Above_200EMA, "0",     "center", False),
        (pct20,            "0.0%",  "center", False),
        (pct50,            "0.0%",  "center", False),
        (pct200,           "0.0%",  "center", False),
        (comp,             "0.0%",  "center", True),
    ]
    for ci, (val, fmt, aln, bld) in enumerate(row_data, 1):
        data_style(ws1.cell(row=ri, column=ci), val, fmt=fmt, align=aln, bold=bld, bg=bg)
    ws1.row_dimensions[ri].height = 17

# ── Conditional colour scale on % cols F-I ───────────────────────────────────
last_row = 3 + len(df_summary)
for col in ["F", "G", "H", "I"]:
    ws1.conditional_formatting.add(
        f"{col}4:{col}{last_row}",
        ColorScaleRule(start_type="min",        start_color="F8696B",
                       mid_type="percentile",   mid_value=50, mid_color="FFEB84",
                       end_type="max",          end_color="63BE7B"))

# freeze panes below header
ws1.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Ticker Detail  (stocks above 20, 50 AND 200 EMA)
# Cols: Sector | Ticker | Close | 20EMA | 50EMA | 200EMA
# Ordered strongest sector first, sector bands highlighted
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Ticker Detail (Above EMA)")
ws2.sheet_view.showGridLines = False

# ── row heights ──────────────────────────────────────────────────────────────
ws2.row_dimensions[1].height = 28
ws2.row_dimensions[2].height = 36

# ── column widths ─────────────────────────────────────────────────────────────
S2_COLS = [("A", 30), ("B", 14), ("C", 16), ("D", 16), ("E", 16), ("F", 16), ("G", 14)]
for col, w in S2_COLS:
    ws2.column_dimensions[col].width = w

# ── Row 1 : Title ─────────────────────────────────────────────────────────────
ws2.merge_cells("A1:G1")
title_style(ws2["A1"],
    "NSE Stocks Trading Above 20 / 50 / 200-Day EMA   ·   Strongest Sector First", 7)

# ── Row 2 : Column headers ────────────────────────────────────────────────────
HDR2 = ["Sector", "Ticker",
        "Current\nClose (\u20b9)", "20-Day\nEMA (\u20b9)", "50-Day\nEMA (\u20b9)", "200-Day\nEMA (\u20b9)",
        "TradingView\nChart"]
for ci, h in enumerate(HDR2, 1):
    hdr_style(ws2.cell(row=2, column=ci), h)

# ── Build detail rows ─────────────────────────────────────────────────────────
sector_order = {s: i for i, s in enumerate(df_summary["Sector"].tolist())}
df_detail_all = df_detail[df_detail["Above_All"]].copy()
df_detail_all["sec_rank"] = df_detail_all["Sector"].map(sector_order)
df_detail_all.sort_values(["sec_rank", "Ticker"], inplace=True)

# assign alternating sector bands (toggle colour each time sector changes)
band   = 0
prev_s = None
band_col = []
for s in df_detail_all["Sector"]:
    if s != prev_s:
        band = 1 - band
        prev_s = s
    band_col.append(band)
df_detail_all["band"] = band_col

BAND_COLOURS = [C_WHITE, C_SEC_BAND]   # alternates per sector group

ri = 3
prev_sector = None
for _, row in df_detail_all.iterrows():
    bg  = BAND_COLOURS[row["band"]]
    new_sec = (row["Sector"] != prev_sector)

    row_data = [
        (row["Sector"],  None,          "left",   new_sec),   # bold on first row of each sector
        (row["Ticker"],  None,          "center", False),
        (row["Last"],    "#,##0.00",    "right",  False),
        (row["EMA20"],   "#,##0.00",    "right",  False),
        (row["EMA50"],   "#,##0.00",    "right",  False),
        (row["EMA200"],  "#,##0.00",    "right",  False),
    ]
    for ci, (val, fmt, aln, bld) in enumerate(row_data, 1):
        data_style(ws2.cell(row=ri, column=ci), val, fmt=fmt, align=aln, bold=bld, bg=bg)

    # TradingView hyperlink in column G
    tv_url  = f"https://www.tradingview.com/chart/?symbol=NSE:{row['Ticker']}"
    tv_cell = ws2.cell(row=ri, column=7)
    tv_cell.value     = f'=HYPERLINK("{tv_url}","Chart 📈")'
    tv_cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")
    tv_cell.fill      = PatternFill("solid", fgColor=bg)
    tv_cell.alignment = Alignment(horizontal="center", vertical="center")
    tv_cell.border    = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"))

    ws2.row_dimensions[ri].height = 16
    prev_sector = row["Sector"]
    ri += 1

ws2.freeze_panes = "A3"

print(f"  Sheet 1 → {len(df_summary)} sectors")
print(f"  Sheet 2 → {len(df_detail_all)} tickers above all three EMAs")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\n✅  Report saved → {OUTPUT_PATH}")

# ── Console summary ───────────────────────────────────────────────────────────
print("\n" + "="*72)
print(f"{'Sector':<30} {'Total':>6}  {'#>20':>5}  {'#>50':>5}  {'#>200':>5}  "
      f"{'%>20':>6}  {'%>50':>6}  {'%>200':>6}  {'Str%':>6}")
print("-"*72)
for _, r in df_summary.iterrows():
    print(f"{r['Sector']:<30} {r['Total_Stocks']:>6}  {r['Above_20EMA']:>5}  "
          f"{r['Above_50EMA']:>5}  {r['Above_200EMA']:>5}  "
          f"{r['Pct_Above_20EMA']:>5.1f}%  {r['Pct_Above_50EMA']:>5.1f}%  "
          f"{r['Pct_Above_200EMA']:>5.1f}%  {r['Strength']:>5.1f}%")
print("="*72)